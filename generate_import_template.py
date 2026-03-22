import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os


CHEMICAL_SUBCATEGORIES = [
    '主胶',
    '树脂',
    '溶剂',
    '助剂',
    '色浆',
    '固化剂'
]

FILM_SUBCATEGORIES = [
    '基材-PET',
    '基材-BOPP',
    '基材-PE',
    '基材-PO',
    '基材-PI',
    '离型膜',
    '保护膜',
    '胶带',
    '硬化膜'
]

CHEMICAL_UNITS = ['kg', 'g', 'L', 'mL']
FILM_UNITS = ['m', 'm²']


def create_template():
    wb = openpyxl.Workbook()
    
    # ==========================
    # 0. 主数据表 (Sheet 1)
    # ==========================
    ws = wb.active
    ws.title = "物料导入表"

    # ==========================
    # 1. 样式预设
    # ==========================
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color="D1D5DB"), right=Side(style='thin', color="D1D5DB"),
        top=Side(style='thin', color="D1D5DB"), bottom=Side(style='thin', color="D1D5DB")
    )

    # ==========================
    # 2. 绘制列头
    # ==========================
    headers = ["产品代码", "物料名称", "类别", "子类别", "默认单位", "供应商", "厂家型号"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    widths = [12, 30, 10, 22, 12, 20, 25]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 只把前 50 行的边框画出来作为视觉体验，否则文件会变大
    for row in range(2, 51):
        ws[f'A{row}'].number_format = '@'
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # ==========================
    # 3. 隐藏配置页
    # ==========================
    config_ws = wb.create_sheet(title="Config")
    config_ws.sheet_state = 'hidden'

    config_data = {
        "化材_子类": CHEMICAL_SUBCATEGORIES,
        "膜材_子类": FILM_SUBCATEGORIES,
        "化材_单位": CHEMICAL_UNITS,
        "膜材_单位": FILM_UNITS
    }

    from openpyxl.workbook.defined_name import DefinedName
    col_idx = 1
    for name, items in config_data.items():
        config_ws.cell(row=1, column=col_idx, value=name)
        for row_idx, item in enumerate(items, 2):
            config_ws.cell(row=row_idx, column=col_idx, value=item)
            
        col_letter = get_column_letter(col_idx)
        ref = f"Config!${col_letter}$2:${col_letter}${len(items) + 1}"
        new_name = DefinedName(name, attr_text=ref)
        wb.defined_names.add(new_name)
        col_idx += 1

    config_ws.protection.sheet = True
    config_ws.protection.enable()
    config_ws.protection.formatColumns = False
    config_ws.protection.formatRows = False
    config_ws.protection.insertColumns = False
    config_ws.protection.insertRows = False

    # ==========================
    # 4. 强力校验规则 (覆盖 3000 行)
    # ==========================
    base_max_row = 3000

    dv_code = DataValidation(type="custom", formula1='AND(ISNUMBER(VALUE(A2)), LEN(A2)=3)')
    dv_code.errorStyle = 'stop'
    dv_code.showErrorMessage = True
    dv_code.showInputMessage = True
    dv_code.errorTitle = '无效的代码格式'
    dv_code.error = '⚠️ 必须且只能输入「刚好3位且必为数字」！不足3位请用0补齐（例：001）。'
    dv_code.promptTitle = '代码规则'
    dv_code.prompt = '请输入3位纯数字代码 (如: 001)'
    ws.add_data_validation(dv_code)
    dv_code.add(f'A2:A{base_max_row}')

    dv_cat = DataValidation(type="list", formula1='"化材,膜材"')
    dv_cat.errorTitle = "输入无效"
    dv_cat.error = "⚠️ 系统只能识别「化材」或「膜材」，请从下拉菜单中选择！"
    ws.add_data_validation(dv_cat)
    dv_cat.add(f'C2:C{base_max_row}')

    dv_subcat = DataValidation(type="list", formula1='INDIRECT($C2&"_子类")')
    dv_subcat.errorTitle = "选择无效"
    dv_subcat.error = "⚠️ 请先在左侧选择正确的大小类别，然后再从本下拉单中挑选子类！"
    ws.add_data_validation(dv_subcat)
    dv_subcat.add(f'D2:D{base_max_row}')

    dv_unit = DataValidation(type="list", formula1='INDIRECT($C2&"_单位")')
    dv_unit.errorTitle = "单位无效"
    dv_unit.error = "⚠️ 请从下拉菜单中选择该类型允许的标准单位！"
    ws.add_data_validation(dv_unit)
    dv_unit.add(f'E2:E{base_max_row}')

    # ==========================
    # 5. 独立的填写说明和示例呈现区 (Sheet 2)
    # ==========================
    ws_help = wb.create_sheet(title="【必看】填写指导与示例")
    
    instructions = [
        "【重要：新手入门指南】",
        "",
        "▶ 极简导出说明（无需再手动删说明了！）：",
        "  相比以前的混乱做法，现在【说明文字】和【表格数据】已经在不同的一页了！",
        "  您只需要在左下角第一张「物料导入表」里安心填数据，",
        "  填完后，不要乱点，直接保持在第一页，点击 Excel 顶部：",
        "  文件 → 另存为 → 选取 (CSV 逗号分隔) 格式，直接存下来即可！",
        "  ⭐️ CSV天生只会保存当前正在看的那一张表，所以这些说明文字会自动被丢弃，完全不会影响后台导入！",
        "",
        "▶ 操作步骤与安全边界：",
        "1. 请点击下方左侧的 Sheet：「物料导入表」开始工作。",
        "2. 从第 2 行开始往下正常录入，中途【绝对不要留任何空白行】。",
        "3. 安全防呆校验规则已经扩展足足覆盖了 3000 行，再多也不怕！",
        "",
        "▶ 各列字段详述：（带 ★ 代表必填项）",
        "★ 产品代码：模板校验要求填写 3 位纯数字（如 001）。系统导入时也能兼容 1 / 01 / J-001 这类输入，但最终都会统一成标准编码。",
        "★ 物料名称：物料对外的正式中文全称。",
        "★ 类别：点开下拉菜单，只能从“化材”和“膜材”里点选。",
        "★ 子类别：它会根据你刚才选的类别联动变化，但只允许选择系统内已启用的正式子类别。",
        "  如果现有子类别不适用，请先由管理员在系统的“子类别管理”里新增正式子类别，再重新填写模板。",
        "  请不要手工修改隐藏的 Config 页来新增子类别，系统不会接受模板里私自扩展的分类。",
        "  默认单位：一样也是联动变化，且只允许化材 kg/g/L/mL，膜材 m/m²。",
        "  供应商 / 厂家型号：全选填。",
        "",
        "▶ 最后给你看一下长啥样的标准示范参考："
    ]

    for i, text in enumerate(instructions, 1):
        cell = ws_help.cell(row=i, column=1, value=text)
        if text.startswith("【重要") or "必看" in text or "黑科技联动" in text:
            cell.font = Font(bold=True, size=12, color="1E3B70")
        elif text.startswith("▶"):
            cell.font = Font(bold=True, size=11, color="374151")
        else:
            cell.font = Font(size=10, color="4B5563")

    ws_help.column_dimensions['A'].width = 90

    # 样例区域
    exam_start = len(instructions) + 1
    exam_headers = ["产品代码", "物料名称", "类别", "子类别", "默认单位", "供应商", "厂家型号"]
    exam_data = [
        ['001', '异丙醇', '化材', '溶剂', 'L', '国药', 'IPA-99'],
        ['002', 'PET保护膜', '膜材', '保护膜', 'm', '东丽', 'T100']
    ]

    for c, val in enumerate(exam_headers, 1):
        cell = ws_help.cell(row=exam_start, column=c, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        ws_help.column_dimensions[get_column_letter(c)].width = widths[c-1]

    for r_offset, row_data in enumerate(exam_data, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws_help.cell(row=exam_start + r_offset, column=c, value=val)
            cell.border = thin_border
            cell.font = Font(color="6B7280")

    # ==========================
    # 6. 保存收尾
    # ==========================
    wb.active = ws  # 保存时永远默认焦距在填报页第一页
    output_path = os.path.join(os.getcwd(), '标准物料导入模板_智能版.xlsx')
    wb.save(output_path)
    print(output_path)

if __name__ == '__main__':
    create_template()
